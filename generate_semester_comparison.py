import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter

SRC = "/Users/meowmeow/Desktop/建德市作业监测数据明细_2025-2026学年_第二学期_a9a2b685eddf4738932a4c8b1d841369.xlsx"
DST = "/Users/meowmeow/Desktop/建德市作业监测_学期对比.xlsx"

COL_NAMES = ["学校", "应用班级数", "应用教师数", "应用作业数",
             "班周均应用次数", "提交作业份数", "提交率", "报告查看率", "讲评率"]

# 0-based column indices for special formatting
PCT_COLS = {5, 6, 7}     # 提交率, 报告查看率, 讲评率 → "xx.xx%"
DECIMAL_COLS = {3}        # 班周均应用次数 → 1 decimal

GREEN = "FF00B050"
RED = "FFFF0000"
BLACK = "FF262626"
GRAY = "FF808080"


def parse_value(val, col_idx):
    if val is None:
        return None
    if col_idx in PCT_COLS:
        s = str(val).strip().replace("%", "")
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def format_value(val, col_idx):
    if val is None:
        return "—"
    if col_idx in PCT_COLS:
        return f"{val:.2f}%"
    if col_idx in DECIMAL_COLS:
        return f"{val:.1f}"
    return str(int(val))


def format_diff(diff, col_idx):
    adiff = abs(diff)
    if col_idx in PCT_COLS:
        return f"{adiff:.2f}%"
    if col_idx in DECIMAL_COLS:
        return f"{adiff:.1f}"
    return str(int(adiff))


def read_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=9, values_only=True):
        school = str(row[0]).strip() if row[0] else None
        if not school:
            continue
        data[school] = [parse_value(row[c], c - 1) for c in range(1, 9)]
    return data


def build_rows(data_s1, data_s2):
    rows = []
    for school in data_s2:
        s2 = data_s2[school]
        s1 = data_s1.get(school)

        # Each cell: (base_str, change_str, change_color)
        cells = []
        for c in range(8):
            v2 = s2[c]
            v1 = s1[c] if s1 else None

            if v2 is None or v1 is None:
                fallback = format_value(v2, c) if v2 is not None else "—"
                cells.append((fallback, " N/A", GRAY))
                continue

            diff = round(v2 - v1, 6)
            base = format_value(v2, c)

            if diff > 0:
                cells.append((base, f" ↑{format_diff(diff, c)}", GREEN))
            elif diff < 0:
                cells.append((base, f" ↓{format_diff(diff, c)}", RED))
            else:
                cells.append((base, " 持平", GRAY))

        rows.append((school, cells))
    return rows


def build_output():
    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    data_s1 = read_sheet(wb_src["第一学期"])
    data_s2 = read_sheet(wb_src["第二学期"])
    wb_src.close()

    rows = build_rows(data_s1, data_s2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学期对比"

    # ── Styles ──
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    title_font = Font(name="宋体", bold=True, size=12)
    title_align = Alignment(horizontal="center", vertical="center")

    header_font = Font(name="楷体", bold=True, size=11)
    header_fill = PatternFill(patternType="solid", fgColor=Color(theme=8, tint=0.6))
    header_align = Alignment(horizontal="center", vertical="center", wrapText=True)

    data_align = Alignment(horizontal="center", vertical="center")

    # ── Row 1: Title ──
    ws.merge_cells("A1:I1")
    ws["A1"] = "建德市初中区本作业应用学期对比（第二学期对比第一学期）"
    ws["A1"].font = title_font
    ws["A1"].alignment = title_align
    ws.row_dimensions[1].height = 15

    # ── Row 2: Headers ──
    for c, name in enumerate(COL_NAMES, 1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 42

    # ── Data rows ──
    for r, (school, cells) in enumerate(rows, 3):
        # Column A: school name
        cell_a = ws.cell(row=r, column=1, value=school)
        cell_a.font = Font(name="Arial", size=10.5, color=BLACK)
        cell_a.alignment = data_align
        cell_a.border = thin_border

        # Columns B–I: metric values with rich text
        for c, (base, change, change_color) in enumerate(cells, 2):
            rich = CellRichText(
                TextBlock(InlineFont(color=BLACK, rFont="Arial", sz=10.5), base),
                TextBlock(InlineFont(color=change_color, rFont="Arial", sz=10.5), change),
            )
            cell = ws.cell(row=r, column=c, value=rich)
            cell.alignment = data_align
            cell.border = thin_border

    # ── Column widths ──
    ws.column_dimensions["A"].width = 27.55
    for letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[letter].width = 15.36

    # ── Freeze header ──
    ws.freeze_panes = "A3"

    wb.save(DST)
    print(f"Done → {DST}")
    print(f"Schools: {len(rows)}")


if __name__ == "__main__":
    build_output()
